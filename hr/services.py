"""
HR Services — Business logic layer for the HR & Staff Management module.
Keeps views thin; all calculations and side effects are encapsulated here.
"""
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from django.db.models import Count, Q
from users.models import AuditLog


# ──────────────────────────────────────────────────────────────────
# Dashboard Statistics Service
# ──────────────────────────────────────────────────────────────────

class HRDashboardService:
    """Aggregates KPI counts and Chart.js datasets for the HR Dashboard."""

    @staticmethod
    def get_stats():
        from .models import Employee, Department
        qs = Employee.objects.all()
        active_qs = qs.filter(employment_status=Employee.EmploymentStatus.ACTIVE)

        stats = {
            'total_employees': qs.count(),
            'active_employees': active_qs.count(),
            'permanent': qs.filter(employment_type=Employee.EmploymentType.PERMANENT).count(),
            'contract': qs.filter(employment_type=Employee.EmploymentType.CONTRACT).count(),
            'consultants': qs.filter(employment_type=Employee.EmploymentType.CONSULTANT).count(),
            'interns': qs.filter(employment_type=Employee.EmploymentType.INTERN).count(),
            'male': qs.filter(gender=Employee.Gender.MALE).count(),
            'female': qs.filter(gender=Employee.Gender.FEMALE).count(),
            'retired': qs.filter(employment_status=Employee.EmploymentStatus.RETIRED).count(),
            'resigned': qs.filter(employment_status=Employee.EmploymentStatus.RESIGNED).count(),
            'on_leave': qs.filter(employment_status=Employee.EmploymentStatus.ON_LEAVE).count(),
            'terminated': qs.filter(employment_status=Employee.EmploymentStatus.TERMINATED).count(),
            'total_departments': Department.objects.filter(is_active=True).count(),
        }
        return stats

    @staticmethod
    def get_gender_chart_data():
        from .models import Employee
        male = Employee.objects.filter(gender=Employee.Gender.MALE).count()
        female = Employee.objects.filter(gender=Employee.Gender.FEMALE).count()
        other = Employee.objects.filter(gender=Employee.Gender.OTHER).count()
        return {
            'labels': ['Male', 'Female', 'Other'],
            'data': [male, female, other],
            'colors': ['#3b82f6', '#ec4899', '#a855f7'],
        }

    @staticmethod
    def get_employment_type_chart_data():
        from .models import Employee
        data = Employee.objects.values('employment_type').annotate(count=Count('id'))
        type_map = dict(Employee.EmploymentType.choices)
        labels = [str(type_map.get(d['employment_type'], d['employment_type'])) for d in data]
        counts = [d['count'] for d in data]
        return {
            'labels': labels,
            'data': counts,
            'colors': ['#10b981', '#f59e0b', '#6366f1', '#14b8a6'],
        }

    @staticmethod
    def get_department_chart_data():
        from .models import Employee
        data = (
            Employee.objects
            .filter(department__isnull=False)
            .values('department__name')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        return {
            'labels': [str(d['department__name']) for d in data],
            'data': [d['count'] for d in data],
        }

    @staticmethod
    def get_staff_level_chart_data():
        from .models import Employee
        data = (
            Employee.objects
            .filter(staff_level__isnull=False)
            .values('staff_level__name', 'staff_level__rank')
            .annotate(count=Count('id'))
            .order_by('staff_level__rank')
        )
        return {
            'labels': [str(d['staff_level__name']) for d in data],
            'data': [d['count'] for d in data],
        }

    @staticmethod
    def get_employment_status_chart_data():
        from .models import Employee
        data = Employee.objects.values('employment_status').annotate(count=Count('id'))
        status_map = dict(Employee.EmploymentStatus.choices)
        labels = [str(status_map.get(d['employment_status'], d['employment_status'])) for d in data]
        counts = [d['count'] for d in data]
        colors = {
            'ACTIVE': '#10b981',
            'ON_LEAVE': '#f59e0b',
            'RETIRED': '#6b7280',
            'RESIGNED': '#3b82f6',
            'TERMINATED': '#ef4444',
        }
        return {
            'labels': labels,
            'data': counts,
            'colors': [colors.get(d['employment_status'], '#9ca3af') for d in data],
        }

    @staticmethod
    def get_age_distribution_chart_data():
        from .models import Employee
        today = timezone.localdate()
        bins = {'Under 25': 0, '25–34': 0, '35–44': 0, '45–54': 0, '55+': 0, 'Unknown': 0}
        for emp in Employee.objects.filter(date_of_birth__isnull=False):
            age = today.year - emp.date_of_birth.year - (
                (today.month, today.day) < (emp.date_of_birth.month, emp.date_of_birth.day)
            )
            if age < 25:
                bins['Under 25'] += 1
            elif age < 35:
                bins['25–34'] += 1
            elif age < 45:
                bins['35–44'] += 1
            elif age < 55:
                bins['45–54'] += 1
            else:
                bins['55+'] += 1
        unknown = Employee.objects.filter(date_of_birth__isnull=True).count()
        bins['Unknown'] = unknown
        return {
            'labels': list(bins.keys()),
            'data': list(bins.values()),
        }

    @staticmethod
    def get_staff_growth_chart_data():
        """Returns cumulative monthly staff count for the past 12 months."""
        from .models import Employee
        today = timezone.localdate()
        months = []
        counts = []
        for i in range(11, -1, -1):
            point = today - relativedelta(months=i)
            label = point.strftime('%b %Y')
            count = Employee.objects.filter(start_date__lte=point).count()
            months.append(label)
            counts.append(count)
        return {'labels': months, 'data': counts}


# ──────────────────────────────────────────────────────────────────
# Contract Monitoring Service
# ──────────────────────────────────────────────────────────────────

class ContractMonitoringService:
    """Filters employees by contract status and expiry windows."""

    @staticmethod
    def _contract_qs():
        from .models import Employee
        return Employee.objects.filter(
            contract_end_date__isnull=False
        ).select_related('department', 'staff_level')

    @classmethod
    def expiring_within(cls, days):
        today = timezone.localdate()
        threshold = today + relativedelta(days=days)
        return cls._contract_qs().filter(
            contract_end_date__gte=today,
            contract_end_date__lte=threshold
        ).order_by('contract_end_date')

    @classmethod
    def expired(cls):
        today = timezone.localdate()
        return cls._contract_qs().filter(contract_end_date__lt=today).order_by('-contract_end_date')

    @classmethod
    def recently_renewed(cls, days=30):
        """Employees whose contract end date was recently updated/extended."""
        today = timezone.localdate()
        threshold = today - relativedelta(days=days)
        return cls._contract_qs().filter(
            updated_at__date__gte=threshold,
            contract_end_date__gte=today
        ).order_by('-updated_at')


# ──────────────────────────────────────────────────────────────────
# HR Report Service
# ──────────────────────────────────────────────────────────────────

class HRReportService:
    """Generates filtered employee querysets and formats for export (CSV, Excel, PDF)."""

    @staticmethod
    def get_filtered_queryset(
        department=None, gender=None,
        employment_type=None, staff_level=None,
        employment_status=None, date_from=None, date_to=None,
        position=None, section=None, education=None, q=None, selected_ids=None
    ):
        from .models import Employee
        qs = Employee.objects.select_related('department', 'section', 'staff_level').prefetch_related('education_records').all()
        if selected_ids:
            # Priority: export selected staff records
            qs = qs.filter(id__in=selected_ids)
        if q and str(q).strip():
            query_str = str(q).strip()
            qs = qs.filter(
                Q(full_name__icontains=query_str) | Q(employee_number__icontains=query_str) |
                Q(position__icontains=query_str) | Q(email__icontains=query_str) |
                Q(section__name__icontains=query_str) |
                Q(education_records__degree__icontains=query_str) |
                Q(education_records__institution__icontains=query_str) |
                Q(education_records__field_of_study__icontains=query_str)
            )
        if department and str(department).strip():
            qs = qs.filter(department_id=department)
        if section and str(section).strip():
            qs = qs.filter(section_id=section)
        if gender and str(gender).strip():
            qs = qs.filter(gender=gender)
        if employment_type and str(employment_type).strip():
            qs = qs.filter(employment_type=employment_type)
        if staff_level and str(staff_level).strip():
            qs = qs.filter(staff_level_id=staff_level)
        if employment_status and str(employment_status).strip():
            qs = qs.filter(employment_status=employment_status)
        if position and str(position).strip():
            qs = qs.filter(position__icontains=position)
        if education and str(education).strip():
            qs = qs.filter(
                Q(education_records__degree__icontains=education) |
                Q(education_records__institution__icontains=education) |
                Q(education_records__field_of_study__icontains=education)
            )
        if date_from and str(date_from).strip():
            qs = qs.filter(start_date__gte=date_from)
        if date_to and str(date_to).strip():
            qs = qs.filter(start_date__lte=date_to)
        return qs.distinct().order_by('department__name', 'full_name')

    @staticmethod
    def get_export_data(employees, data_section, department_id=None, export_fields=None):
        """Return only the selected columns and rows for one export table.

        Education intentionally remains one row per qualification, so every
        qualification is retained when a staff member has several records.
        """
        from .models import Department, DepartmentSection, EmployeeDocument, EmployeeEducation
        export_fields = set(export_fields or [])

        def selected_columns(columns):
            # Backwards-compatible exports without field parameters include all
            # columns. Once fields are supplied, a table contains only its own
            # selected fields.
            if not export_fields:
                return columns
            return [column for column in columns if f'{data_section}:{column[0]}' in export_fields]

        if data_section == 'staff_directory':
            queryset = employees

            def education_summary(employee):
                records = employee.education_records.all()
                return '; '.join(
                    ' — '.join(str(value) for value in (
                        record.degree,
                        record.institution,
                        record.field_of_study or None,
                        record.year_completed or None,
                    ) if value)
                    for record in records
                )

            columns = selected_columns([
                ('employee_number', 'Employee Number', lambda item: item.employee_number),
                ('full_name', 'Name', lambda item: item.full_name),
                ('department', 'Department', lambda item: item.department.name if item.department else ''),
                ('section', 'Section', lambda item: item.section.name if item.section else ''),
                ('position', 'Position', lambda item: item.position),
                ('education_summary', 'Education', education_summary),
                ('employment_type', 'Employment Type', lambda item: item.get_employment_type_display()),
                ('staff_level', 'Staff Level', lambda item: str(item.staff_level) if item.staff_level else ''),
                ('employment_status', 'Employment Status', lambda item: item.get_employment_status_display()),
                ('phone', 'Phone', lambda item: item.phone),
                ('email', 'Email', lambda item: item.email),
            ])
            title = 'Staff Directory'
        elif data_section == 'education_information':
            queryset = EmployeeEducation.objects.select_related('employee__department').filter(employee__in=employees)
            columns = selected_columns([
                ('employee_number', 'Employee Number', lambda item: item.employee.employee_number),
                ('employee_name', 'Name', lambda item: item.employee.full_name),
                ('department', 'Department', lambda item: item.employee.department.name if item.employee.department else ''),
                ('degree', 'Degree / Qualification', lambda item: item.degree),
                ('institution', 'Institution', lambda item: item.institution),
                ('field_of_study', 'Field of Study', lambda item: item.field_of_study),
                ('year_completed', 'Year Completed', lambda item: item.year_completed or ''),
            ])
            title = 'Education Records'
        elif data_section == 'staff_documents':
            queryset = EmployeeDocument.objects.select_related('employee').filter(employee__in=employees)
            columns = selected_columns([
                ('employee_number', 'Employee Number', lambda item: item.employee.employee_number),
                ('employee_name', 'Name', lambda item: item.employee.full_name),
                ('title', 'Document Title', lambda item: item.title),
                ('document_type', 'Document Type', lambda item: item.get_document_type_display()),
                ('uploaded_at', 'Uploaded Date', lambda item: item.uploaded_at.date()),
            ])
            title = 'Staff Documents'
        elif data_section == 'departments':
            queryset = Department.objects.select_related('head').all()
            if department_id:
                queryset = queryset.filter(pk=department_id)
            columns = selected_columns([
                ('name', 'Department', lambda item: item.name),
                ('code', 'Code', lambda item: item.code),
                ('head', 'Head of Department', lambda item: item.head.full_name if item.head else ''),
                ('active_sections', 'Active Sections', lambda item: item.section_count),
                ('status', 'Status', lambda item: 'Active' if item.is_active else 'Inactive'),
            ])
            title = 'Departments'
        elif data_section == 'sections':
            queryset = DepartmentSection.objects.select_related('department').all()
            if department_id:
                queryset = queryset.filter(department_id=department_id)
            columns = selected_columns([
                ('department', 'Department', lambda item: item.department.name),
                ('name', 'Section', lambda item: item.name),
                ('code', 'Code', lambda item: item.code),
                ('description', 'Description', lambda item: item.description),
                ('order', 'Display Order', lambda item: item.order),
                ('status', 'Status', lambda item: 'Active' if item.is_active else 'Inactive'),
            ])
            title = 'Department Sections'
        elif data_section == 'employment_details':
            queryset = employees
            columns = selected_columns([
                ('employee_number', 'Employee Number', lambda item: item.employee_number),
                ('full_name', 'Name', lambda item: item.full_name),
                ('department', 'Department', lambda item: item.department.name if item.department else ''),
                ('section', 'Section', lambda item: item.section.name if item.section else ''),
                ('position', 'Position', lambda item: item.position),
                ('employment_type', 'Employment Type', lambda item: item.get_employment_type_display()),
                ('staff_level', 'Staff Level', lambda item: str(item.staff_level) if item.staff_level else ''),
                ('start_date', 'Start Date', lambda item: item.start_date),
                ('contract_end_date', 'Contract End Date', lambda item: item.contract_end_date or ''),
                ('employment_status', 'Employment Status', lambda item: item.get_employment_status_display()),
            ])
            title = 'Employment Details'
        else:
            queryset = employees
            columns = selected_columns([
                ('employee_number', 'Employee Number', lambda item: item.employee_number),
                ('full_name', 'Name', lambda item: item.full_name),
                ('gender', 'Gender', lambda item: item.get_gender_display()),
                ('date_of_birth', 'Date of Birth', lambda item: item.date_of_birth or ''),
                ('nationality', 'Nationality', lambda item: item.nationality),
                ('phone', 'Phone', lambda item: item.phone),
                ('email', 'Email', lambda item: item.email),
            ])
            title = 'Staff Information'

        if not columns:
            return title, [], []
        return title, [label for _key, label, _value in columns], [
            [value(item) for _key, _label, value in columns] for item in queryset
        ]

    @classmethod
    def export_csv(cls, employees, response, data_sections, department_id=None, export_fields=None):
        import csv
        writer = csv.writer(response)
        written_sections = 0
        for data_section in data_sections:
            title, headers, rows = cls.get_export_data(employees, data_section, department_id, export_fields)
            if not headers:
                continue
            if written_sections:
                writer.writerow([])
            writer.writerow([title])
            writer.writerow(headers)
            writer.writerows(rows)
            written_sections += 1

    @classmethod
    def export_excel(cls, employees, response, data_sections, department_id=None, export_fields=None):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export.")

        workbook = openpyxl.Workbook()
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        worksheet_count = 0
        for data_section in data_sections:
            title, headers, rows = cls.get_export_data(employees, data_section, department_id, export_fields)
            if not headers:
                continue
            worksheet = workbook.active if worksheet_count == 0 else workbook.create_sheet()
            worksheet_count += 1
            worksheet.title = title[:31]
            for column, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=column, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for row in rows:
                worksheet.append(row)
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 40)
        workbook.save(response)

    @classmethod
    def export_pdf(cls, employees, response, data_sections, department_id=None, export_fields=None):
        """Generate a compact table PDF for the selected HR data section."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
        except ImportError:
            raise ImportError("reportlab is required for PDF export.")

        document = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=18, leftMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('ExportTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1E3A5F'), spaceAfter=4)
        subtitle_style = ParagraphStyle('ExportSubtitle', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#64748B'), spaceAfter=12)
        cell_style = ParagraphStyle('ExportCell', parent=styles['Normal'], fontSize=7, leading=9)
        header_style = ParagraphStyle('ExportHeader', parent=cell_style, textColor=colors.white, fontName='Helvetica-Bold')
        elements = []
        for data_section in data_sections:
            title, headers, rows = cls.get_export_data(employees, data_section, department_id, export_fields)
            if not headers:
                continue
            table_rows = [[Paragraph(str(header), header_style) for header in headers]]
            table_rows.extend([[Paragraph(str(value), cell_style) for value in row] for row in rows])
            width = (landscape(A4)[0] - 36) / max(len(headers), 1)
            table = Table(table_rows, colWidths=[width] * len(headers), repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ]))
            elements.extend([
                Paragraph(f"DNMG – {title}", title_style),
                Paragraph(f"Generated on {timezone.localdate():%Y-%m-%d} | Records: {len(rows)}", subtitle_style), table,
            ])
        document.build(elements)


# ──────────────────────────────────────────────────────────────────
# HR Audit Service
# ──────────────────────────────────────────────────────────────────

class HRAuditService:
    """Logs HR actions to the central AuditLog model."""

    @staticmethod
    def log(request, action: str, details: dict = None):
        ip = request.META.get('REMOTE_ADDR')
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            ip_address=ip,
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            details=details or {},
        )
